import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy import select

from app.core.database import SessionLocal, engine
from app.core.database import Base
from app.modules.rutas.models import Color, Linea, LineaPunto, LineaRuta, Punto, Ruta, Trasbordo

XLS_PATH = Path(__file__).resolve().parent.parent / "DatosLineas.xlsx"


def read_xls():
    wb = openpyxl.load_workbook(str(XLS_PATH), data_only=True)
    return wb


def recreate_tables():
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    print("Tablas recreadas correctamente.")


def seed_colores(wb, db):
    sheet = wb["Lineas"]
    colores_vistos: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cod_hex = str(row[2]).strip()
        if cod_hex not in colores_vistos:
            colores_vistos[cod_hex] = cod_hex

    for cod_hex in colores_vistos:
        color = Color(nombre=cod_hex, cod_hex=cod_hex)
        db.add(color)
    db.commit()

    result = db.execute(select(Color))
    return {c.cod_hex: c.id for c in result.scalars().all()}


def seed_puntos(wb, db):
    sheet = wb["Puntos"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stop_val = str(row[4]).strip().upper() if row[4] is not None else "N"
        punto = Punto(
            id=int(row[0]),
            latitud=row[1],
            longitud=row[2],
            descripcion=str(row[3]).strip(),
            stop=(stop_val == "S"),
        )
        db.add(punto)
    db.commit()
    total = sheet.max_row - 1
    print(f"Puntos insertados: {total}")


def seed_rutas(db):
    ruta1 = Ruta(id=1, descripcion="Salida")
    ruta2 = Ruta(id=2, descripcion="Retorno")
    db.add_all([ruta1, ruta2])
    db.commit()
    print("Rutas insertadas: 2")


def seed_lineas(wb, db, color_map):
    sheet = wb["Lineas"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cod_hex = str(row[2]).strip()
        linea = Linea(
            id=int(row[0]),
            nombre=str(row[1]).strip(),
            imagen_url=str(row[3]).strip() if row[3] else None,
            fecha_creada=row[4],
            color_id=color_map[cod_hex],
        )
        db.add(linea)
    db.commit()
    total = sheet.max_row - 1
    print(f"Líneas insertadas: {total}")


def seed_lineas_rutas(wb, db):
    sheet = wb["LineaRuta"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        lr = LineaRuta(
            id=int(row[0]),
            linea_id=int(row[1]),
            ruta_id=int(row[2]),
            descripcion=str(row[3]).strip() if row[3] else None,
            distancia=float(row[4]) if row[4] is not None else 0.0,
            tiempo=float(row[5]) if row[5] is not None else 0.0,
        )
        db.add(lr)
    db.commit()
    total = sheet.max_row - 1
    print(f"LineaRuta insertadas: {total}")


def seed_lineas_puntos(wb, db):
    sheet = wb["LineasPuntos"]
    batch = []
    BATCH_SIZE = 200
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        lp = LineaPunto(
            id=int(row[0]),
            linea_ruta_id=int(row[1]),
            punto_id=int(row[2]),
            punto_destino_id=int(row[3]) if row[3] is not None and int(row[3]) > 0 else None,
            orden=int(row[4]),
            distancia=float(row[5]) if row[5] is not None else 0.0,
            tiempo=float(row[6]) if row[6] is not None else 0.0,
        )
        batch.append(lp)
        if len(batch) >= BATCH_SIZE:
            db.add_all(batch)
            db.commit()
            total += len(batch)
            batch = []
    if batch:
        db.add_all(batch)
        db.commit()
        total += len(batch)
    print(f"LineasPuntos insertados: {total}")


def seed_trasbordos(wb, db):
    sheet = wb["PuntosTrasbordos"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        trasbordo = Trasbordo(
            id=int(row[0]),
            punto_id=int(row[1]),
            linea_origen_id=int(row[2]),
            linea_destino_id=int(row[3]),
            penalizacion_min=float(row[4]),
        )
        db.add(trasbordo)
    db.commit()
    total = sheet.max_row - 1
    print(f"Trasbordos insertados: {total}")


def main():
    wb = read_xls()
    recreate_tables()

    with SessionLocal() as db:
        color_map = seed_colores(wb, db)
        seed_rutas(db)
        seed_puntos(wb, db)
        seed_lineas(wb, db, color_map)
        seed_lineas_rutas(wb, db)
        seed_lineas_puntos(wb, db)
        seed_trasbordos(wb, db)

    print("¡Seed completado exitosamente!")


if __name__ == "__main__":
    main()
