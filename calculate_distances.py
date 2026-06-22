import math
from pathlib import Path

import openpyxl

XLS_PATH = Path(__file__).resolve().parent / "DatosLineas.xlsx"

AVG_SPEED_KMH = 20.0


def haversine(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6371.0
    d_lat = math.radians(lat2 - lat1)
    d_lng = math.radians(lng2 - lng1)
    a = math.sin(d_lat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lng / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def main():
    wb = openpyxl.load_workbook(str(XLS_PATH), data_only=True)

    ws_puntos = wb["Puntos"]
    coords: dict[int, tuple[float, float]] = {}
    for row in ws_puntos.iter_rows(min_row=2, values_only=True):
        pid = int(row[0])
        coords[pid] = (float(row[1]), float(row[2]))

    ws_lp = wb["LineasPuntos"]
    totales_por_ruta: dict[int, float] = {}
    totales_t_por_ruta: dict[int, float] = {}
    for row in ws_lp.iter_rows(min_row=2):
        p_id = int(row[2].value)
        pd_val = row[3].value
        pd_id = None
        if pd_val is not None:
            try:
                pd_int = int(float(pd_val))
                if pd_int > 0:
                    pd_id = pd_int
            except (ValueError, TypeError):
                pass

        lr_id = int(row[1].value)
        dist = 0.0
        tiempo = 0.0
        if pd_id is not None and p_id in coords and pd_id in coords:
            dist = haversine(coords[p_id][0], coords[p_id][1], coords[pd_id][0], coords[pd_id][1])
            tiempo = (dist / AVG_SPEED_KMH) * 60

        row[5].value = round(dist, 6)
        row[6].value = round(tiempo, 6)

        totales_por_ruta[lr_id] = totales_por_ruta.get(lr_id, 0.0) + dist
        totales_t_por_ruta[lr_id] = totales_t_por_ruta.get(lr_id, 0.0) + tiempo

    ws_lr = wb["LineaRuta"]
    for row in ws_lr.iter_rows(min_row=2):
        lr_id = int(row[0].value)
        dist_total = totales_por_ruta.get(lr_id, 0.0)
        tiempo_total = totales_t_por_ruta.get(lr_id, 0.0)
        row[4].value = round(dist_total, 6)
        row[5].value = round(tiempo_total, 6)

    wb.save(str(XLS_PATH))
    total_lp = ws_lp.max_row - 1
    total_lr = ws_lr.max_row - 1
    print(f"Distancia (km) y tiempo (min) calculados para {total_lp} segmentos y {total_lr} rutas.")
    print(f"Velocidad promedio: {AVG_SPEED_KMH} km/h")


if __name__ == "__main__":
    main()
